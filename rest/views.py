from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, HttpResponseRedirect, Http404

#data
from .models import Table3, Table2, Table1, UserLog
import datetime
from django.utils.dateparse import parse_duration, parse_date, parse_time, parse_datetime
from django.utils import timezone

#message
from django.contrib import messages
#forms
from .forms import Table1Form, Table2Form, Table3Form

#user
from django.contrib.auth.models import User, Group
from django.contrib.auth.decorators import  login_required, permission_required
from django.views.decorators.http import require_GET, require_POST
import re
import string

#email
from django.core.mail import send_mail
from django.conf import settings

#auth
from django.contrib.auth import authenticate, login, logout

# Django ORM imports
from django.db.models import Max, Count, Q, Min, Avg, F

#pdf / excel
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from openpyxl import Workbook
from django.utils.timezone import is_aware
from openpyxl.utils import get_column_letter

###### 
# Used to test the error 403
from django.core.exceptions import PermissionDenied

@login_required
def test_403(request):
    """
    Shall remove this in production
    """
    raise PermissionDenied("This is a test 403 error")

# Used to test the error 400, 404, 500
@login_required
def test_400(request):
    return HttpResponse("Bad Request - Invalid parameters", status=400)


@login_required
def test_404(request):
    raise Http404("This page does not exist - test 404 error")

@login_required
def test_500(request):
    raise Exception("This is a test 500 internal server error")

# index view
@require_GET
def index(request):
    return render(request, 'index.html')

# basic redirect to login
@require_GET
def rest_basic(request):
    return redirect('login_rest_basic')

# home view after login
@require_GET
def home_rest_basic(request):
    return render(request, 'home_rest_basic.html')

# User registration view
def user_register(request):
    if request.method == 'POST':
        # Extract and validate form data
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')

        # Basic validation
        if password1 != password2:
            messages.error(request, "Passwords do not match.")
            return render(request, 'RB/register.html')
        
        # Password strength validation
        if len(password1) < 8 or len(password1) > 14:
            messages.error(request, "Password must be between 8 and 14 characters.")
            return render(request, 'RB/register.html')
        
        # Check for letters, numbers, and special characters
        special_characters = string.punctuation
        if not re.search(r'[A-Za-z]', password1) or not re.search(r'\d', password1) or not any(char in special_characters for char in password1):
            messages.error(request, "Password must include letters, numbers, and special characters.")
            return render(request, 'RB/register.html')

        # Check if username or email already exists
        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return render(request, 'RB/register.html')

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email is already in use.")
            return render(request, 'RB/register.html')
        
        # Create the user and assign to the Customers group
        user = User.objects.create_user(username=username, password=password1, email=email)
        viewer_group = Group.objects.get(name='Customers')
        user.groups.add(viewer_group)
        
        messages.success(request, "User registered successfully.")
        
        return redirect('login_rest_basic') 
        
    return render(request, 'RB/register.html')

# User login view
def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        # Basic validation
        if not username or not password:
            messages.error(request, "Please provide both username and password.")
            return render(request, 'RB/login.html')
        
        # Authenticate user
        user = authenticate(request, username=username, password=password)
        
        # If authentication is successful, log the user in
        if user is not None:
            login(request, user)
            # Save current app in session
            request.session['current_app'] = 'rest'
            request.session['home_url'] = 'home_rest_basic'
            request.session['profile_url'] = 'profile_rest_basic'
            request.session['logout_url'] = 'logout_rest_basic'
            return redirect('home_rest_basic')
        else:
            # Invalid credentials
            messages.error(request, "Invalid username or password.")
            return render(request, 'RB/login.html')
    
    # If GET request, just render the login page
    logout_msg = request.session.pop('logout_message', None)
    if logout_msg:
        messages.warning(request, logout_msg)
    return render(request, 'RB/login.html')

# User profile view
@login_required
def profile(request):
    user = request.user
    
    # Determine user role
    if user.groups.filter(name='Admins').exists():
        role = 'Administrator'
    elif user.groups.filter(name='Customers').exists():
        role = 'Customer'
    else:
        role = 'No role assigned'
    
    # Get user permissions
    user_permissions = []
    for group in user.groups.all():
        for permission in group.permissions.all():
            user_permissions.append(permission.name)
    
    # Remove duplicates
    context = {
        'user': user,
        'username': user.username,
        'email': user.email,
        'role': role,
        'permissions': user_permissions,
    }
    return render(request, 'profile.html', context)

# User logout view
@login_required
def user_logout(request):
    if request.method == 'POST':
        logout(request)
        messages.success(request, f"Session closed successfully.")
        return redirect('login_rest_basic')

# If GET request, show confirmation page
@require_GET
def crud(request):
    return render(request, 'crud.html')

# View to display data from all three tables
@require_GET
@permission_required('rest_basic.view_data', raise_exception=True)
def get_data(request):
    table3 = Table3.objects.all()
    table2 = Table2.objects.all()
    table1 = Table1.objects.all()
    return render(request, 'get_data.html',{"table3":table3,"table2":table2,"table1":table1})

# View to add data to the tables
@permission_required('rest_basic.add_data', raise_exception=True)
def add_data(request):
    if request.method == 'POST':
        form_type = request.POST.get('form_type')

        # --- Table3 ---
        if form_type == 'table3':
            duration = request.POST.get('duration_field')
            email = request.POST.get('email_field')

            try:
                # Convert to timedelta if needed
                days, time_part = duration.strip().split(' ')
                hours, minutes, seconds = map(int, time_part.split(':'))
                duration_td = datetime.timedelta(days=int(days), hours=hours, minutes=minutes, seconds=seconds)

                Table3.objects.create(duration_field=duration_td, email_field=email)
                messages.success(request, "Table3 created successfully.")
                return redirect('add_data')
            except Exception as e:
                # Handle invalid duration format or other errors
                messages.error(request, f"Error creating Table3: {e}")
                return redirect('add_data')

        # --- Table2 ---
        elif form_type == 'table2':
            choice = request.POST.get('positive_small_int')
            # Validate choice
            try:
                Table2.objects.create(positive_small_int=int(choice))
                messages.success(request, "Table2 created successfully.")
                return redirect('add_data')
            except Exception as e:
                messages.error(request, f"Error creating Table2: {e}")
                return redirect('add_data')

        # --- Table1 ---
        elif form_type == 'table1':
            try:
                # Django doesn't accept empty string ('') as valid value for fields that allow null=True or blank=True. In those cases, you need to pass None.
                foreign_key_id = request.POST.get('foreign_key') or None
                one_to_one_id = request.POST.get('one_to_one') or None
                many_to_many_ids = request.POST.getlist('many_to_many')
                
                # Handle date field
                date_field_value = request.POST.get('date_field')
                parsed_date = None
                if date_field_value and date_field_value.strip():
                    parsed_date = parse_date(date_field_value.strip())
                
                # Handle time field
                time_field_value = request.POST.get('time_field')
                parsed_time = None
                if time_field_value and time_field_value.strip():
                    parsed_time = parse_time(time_field_value.strip())
                
                # Handle datetime field with timezone awareness
                datetime_field_value = request.POST.get('datetime_field')
                parsed_datetime = None
                if datetime_field_value and datetime_field_value.strip():
                    parsed_datetime = parse_datetime(datetime_field_value.strip())
                    if parsed_datetime and timezone.is_naive(parsed_datetime):
                        parsed_datetime = timezone.make_aware(parsed_datetime)
                
                # Create Table1 instance
                table1 = Table1.objects.create(
                    foreign_key_id=foreign_key_id or None,
                    one_to_one_id=one_to_one_id or None,
                    integer_field=request.POST.get('integer_field') or None,
                    float_field=request.POST.get('float_field') or None,
                    char_field=request.POST.get('char_field'),
                    text_field=request.POST.get('text_field', ''),
                    boolean_field=bool(request.POST.get('boolean_field')),
                    date_field=parsed_date,
                    time_field=parsed_time,
                    datetime_field=parsed_datetime,
                    image_field=request.FILES.get('image_field'),
                    file_field=request.FILES.get('file_field'),
                )
                # Many-to-many linking
                if many_to_many_ids:
                    table3_objs = Table3.objects.filter(id__in=filter(None, many_to_many_ids))
                    table1.many_to_many.set(table3_objs)

                messages.success(request, "Table1 created successfully.")
                return redirect('add_data')
            except Exception as e:
                messages.error(request, f"Error creating Table1: {e}")
                return redirect('add_data')
    else:
        # values list for foreign key and many to many fields      
        table2 = Table2.objects.values_list('id', flat=True)
        table3 = Table3.objects.values_list('id', flat=True)
        # this is for not harcoding the options on the table 2
        table2_choices = Table2._meta.get_field('positive_small_int').choices
        return render(request, 'add_data.html',{
            "table3": table3,
            "table2": table2,
            "table2_choices": table2_choices,
        })

# View to update data in the tables
@permission_required('rest_basic.change_data', raise_exception=True)
def update_data(request):
    # Fetch all records from the three tables
    table1 = Table1.objects.all()
    table2 = Table2.objects.all()
    table3 = Table3.objects.all()
    table2_ids = Table2.objects.values_list('id', flat=True)
    table3_ids = Table3.objects.values_list('id', flat=True)

    editing = None
    editing_table = None
    selected_many = []

    # Post request to save changes
    if request.method == 'POST' and request.POST.get('edit_id') and request.POST.get('edit_table'):
        pk = request.POST.get('edit_id')
        edit_table = request.POST.get('edit_table')

        try:
            # Update the corresponding table based on edit_table value
            if edit_table == 'table1':
                editing = get_object_or_404(Table1, pk=pk)
                editing_table = 'table1'
                editing.foreign_key_id = request.POST.get('foreign_key') or None
                editing.one_to_one_id = request.POST.get('one_to_one') or None
                editing.integer_field = request.POST.get('integer_field') or None
                editing.float_field = request.POST.get('float_field') or None
                editing.char_field = request.POST.get('char_field')
                editing.text_field = request.POST.get('text_field', '')
                editing.boolean_field = bool(request.POST.get('boolean_field'))
                
                # Handle date field
                date_field_value = request.POST.get('date_field')
                if date_field_value:
                    editing.date_field = parse_date(date_field_value)
                else:
                    editing.date_field = None
                    
                # Handle time field
                time_field_value = request.POST.get('time_field')
                if time_field_value:
                    editing.time_field = parse_time(time_field_value)
                else:
                    editing.time_field = None
                    
                # Handle datetime field with timezone awareness
                datetime_field_value = request.POST.get('datetime_field')
                if datetime_field_value:
                    parsed_datetime = parse_datetime(datetime_field_value)
                    if parsed_datetime:
                        # If the datetime is naive, make it timezone-aware
                        if timezone.is_naive(parsed_datetime):
                            parsed_datetime = timezone.make_aware(parsed_datetime)
                        editing.datetime_field = parsed_datetime
                    else:
                        editing.datetime_field = None
                else:
                    editing.datetime_field = None

                # Handle file fields
                if request.FILES.get('image_field'):
                    editing.image_field = request.FILES.get('image_field')
                if request.FILES.get('file_field'):
                    editing.file_field = request.FILES.get('file_field')
                
                editing.save()
                
                # ManyToMany
                many_to_many_ids = request.POST.getlist('many_to_many')
                if many_to_many_ids:
                    table3_objs = Table3.objects.filter(id__in=filter(None, many_to_many_ids))
                    editing.many_to_many.set(table3_objs)
                else:
                    editing.many_to_many.clear()
                messages.success(request, "Table1 updated successfully.")
            elif edit_table == 'table2':
                editing = get_object_or_404(Table2, pk=pk)
                editing_table = 'table2'
                editing.positive_small_int = request.POST.get('positive_small_int')
                editing.save()
                messages.success(request, "Table2 updated successfully.")
            elif edit_table == 'table3':
                editing = get_object_or_404(Table3, pk=pk)
                editing_table = 'table3'
                editing.duration_field = parse_duration(request.POST.get('duration_field'))
                editing.email_field = request.POST.get('email_field')
                editing.save()
                messages.success(request, "Table3 updated successfully.")
            return redirect('update_data')
        except Exception as e:
            messages.error(request, f"Error updating: {e}")
            # Apply PRG pattern on error to avoid browser 'Confirm form resubmission'
            return redirect('update_data')

    # Get request to load the record to edit
    elif request.GET.get('edit_id') and request.GET.get('edit_table'):
        pk = request.GET.get('edit_id')
        editing_table = request.GET.get('edit_table')
        if editing_table == 'table1':
            editing = get_object_or_404(Table1, pk=pk)
            selected_many = list(editing.many_to_many.values_list('id', flat=True))
        elif editing_table == 'table2':
            editing = get_object_or_404(Table2, pk=pk)
        elif editing_table == 'table3':
            editing = get_object_or_404(Table3, pk=pk)

    return render(request, 'update_data.html', {
        "table1": table1,
        "table2": table2,
        "table3": table3,
        "editing": editing,
        "editing_table": editing_table,
        "selected_many": selected_many,
        "table2_ids": table2_ids,
        "table3_ids": table3_ids,
        "table2_choices": Table2._meta.get_field('positive_small_int').choices,
    })

# View to delete data from Table1 (soft delete by marking as inactive)
@permission_required('rest_basic.change_data', raise_exception=True)
def delete_data_1(request):
    active_items = Table1.objects.filter(boolean_field=True)
    inactive_items = Table1.objects.filter(boolean_field=False)
    deleting = None

    # If a POST request is received with a delete_id, set the record as inactive
    if request.method == 'POST' and request.POST.get('delete_id'):
        pk = request.POST.get('delete_id')
        entry = get_object_or_404(Table1, pk=pk)
        entry.boolean_field = False  # Mark as inactive
        entry.save()
        messages.success(request, "Record disabled successfully.")
        return redirect('delete_data_1')

    # If a GET request is received with a delete_id, show the confirmation prompt
    elif request.GET.get('delete_id'):
        pk = request.GET.get('delete_id')
        deleting = get_object_or_404(Table1, pk=pk)

    # Render the template with both active and inactive items and the item to confirm disabling
    return render(request, 'delete_data_1.html', {
        "active_items": active_items,
        "inactive_items": inactive_items,
        "deleting": deleting,
    })

# View to permanently delete data from Table1
@permission_required('rest_basic.delete_data', raise_exception=True)
def delete_data_2(request):
    records = Table1.objects.all()
    deleting = None

    if request.method == 'POST' and request.POST.get('delete_id'):
        pk = request.POST.get('delete_id')
        entry = get_object_or_404(Table1, pk=pk)
        # Delete associated files if they exist
        if entry.image_field:
            entry.image_field.delete(save=False)
        if entry.file_field:
            entry.file_field.delete(save=False)
        entry.delete()
        messages.success(request, "Record permanently deleted.")
        return redirect('delete_data_2')
    # If an ID is provided in the request "GET", shows the confirmation
    elif request.GET.get('delete_id'):
        pk = request.GET.get('delete_id')
        deleting = get_object_or_404(Table1, pk=pk)

    return render(request, 'delete_data_2.html', {
        "records": records,
        "deleting": deleting,
    })

# Form-based CRUD views
#@require_GET
#def crud_form(request):
#    return render(request, 'crud_form.html')

# View to display data from all three tables using forms
@require_GET
@permission_required('rest_basic.view_data', raise_exception=True)
def get_data_form(request):
    table3 = Table3.objects.all()
    table2 = Table2.objects.all()
    table1 = Table1.objects.all()
    return render(request, 'get_data_form.html',{"table3":table3,"table2":table2,"table1":table1})

# View to handle form submissions for adding new records
@permission_required('rest_basic.add_data', raise_exception=True)
def form(request,table):
    form_map = {
        "table1": (Table1Form, "Table1"),
        "table2": (Table2Form, "Table2"),
        "table3": (Table3Form, "Table3"),
    }
    # Get the appropriate form class based on the table parameter
    form_class, model_name = form_map.get(table, (None, None))
    if not form_class:
        return HttpResponse("Table is not valid", status=400)
    
    # Handle form submission
    if request.method == 'POST':
        form = form_class(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, f"{model_name} created successfully.")
            return redirect(get_data_form)
        messages.error(request, "Invalid form.")
    else:
        form = form_class()

    return render(request, 'form.html', {'form': form, 'model_name': model_name})

# View to display the form selection page for adding new records
@require_GET
@permission_required('rest_basic.add_data', raise_exception=True)
def add_data_form(request):
    return render(request, 'add_data_form.html')

# View to handle form submissions for updating existing records
@permission_required('rest_basic.change_data', raise_exception=True)
def update_form(request,table,id):
    form_class = None
    model_name = ""
    
    # Determine the form class and model name based on the table parameter
    if table == "table1":
        form_class = Table1Form
        model_name = "Table1"
        obj = get_object_or_404(Table1, pk=id)
    elif table == "table2":
        form_class = Table2Form
        model_name = "Table2"
        obj = get_object_or_404(Table2, pk=id)
    elif table == "table3":
        form_class = Table3Form
        model_name = "Table3"
        obj = get_object_or_404(Table3, pk=id)
    else:
        return HttpResponse("Table is not valid", status=400)

    # Handle form submission
    if request.method == 'POST':
        form = form_class(request.POST, request.FILES, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, f"{model_name} updated successfully.")
            return redirect(update_data_form)
        else:
            messages.error(request, "Invalid form.")
    else:
        form = form_class(instance=obj)

    return render(request, 'form.html', {'form': form, 'model_name': model_name})

# View to display the form selection page for updating records
@require_GET
@permission_required('rest_basic.view_data', raise_exception=True)
def update_data_form(request):
    table3 = Table3.objects.all()
    table2 = Table2.objects.all()
    table1 = Table1.objects.all()
    return render(request, 'update_data_form.html',{"table3":table3,"table2":table2,"table1":table1})

# User management view (only for admins)
@login_required
@permission_required('rest_basic.manage_users', raise_exception=True)
def user_management(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        new_role = request.POST.get('role')
        
        try:
            target_user = User.objects.get(id=user_id)
            
            # Remove from all groups
            target_user.groups.clear()
            
            # Assign new group
            if new_role == 'admin':
                admin_group = Group.objects.get(name='Admins')
                target_user.groups.add(admin_group)
            elif new_role == 'customer':
                customer_group = Group.objects.get(name='Customers')
                target_user.groups.add(customer_group)
            
            messages.success(request, f"Permissions updated for {target_user.username}")
        except User.DoesNotExist:
            messages.error(request, "User not found.")
        except Exception as e:
            messages.error(request, f"Error updating permissions: {e}")
        return redirect('user_management')
    
    # Get all users with their roles
    users_with_roles = []
    for user in User.objects.all():
        if user.groups.filter(name='Admins').exists():
            role = 'Administrator'
        elif user.groups.filter(name='Customers').exists():
            role = 'Customer'
        else:
            role = 'No role'
        
        users_with_roles.append({
            'user': user,
            'role': role
        })
    
    return render(request, 'user_management.html', {
        'users_with_roles': users_with_roles
    })

# View to display user activity logs (only for admins)
@login_required
def user_logs(request):
    logs = UserLog.objects.filter().order_by('-timestamp')[:15]
    return render(request, 'user_logs.html', {
        'logs': logs
    })

# Views for making queries examples
def making_queries(request):
    return render(request, 'making_queries.html')

# View to demonstrate retrieving all records from multiple tables
def all_example(request):
    all_table1 = Table1.objects.all()
    all_table2 = Table2.objects.all()
    all_table3 = Table3.objects.all()
    
    # Create a combined context for all tables
    context = {
        'class': 'Basic',
        'type': 'All',
        'table1_data': all_table1,
        'table2_data': all_table2,
        'table3_data': all_table3,
        'table1_count': all_table1.count(),
        'table2_count': all_table2.count(),
        'table3_count': all_table3.count(),
    }
    
    return render(request, 'all_queries.html', context)

# View to demonstrate filter lookups
def filter_example(request):
    #Example of filter lookups:
    # __in = in
    # __contains = contains
    # __icontains = contains (case insensitive)
    # __startswith = starts with
    # __istartswith = starts with (case insensitive)
    # __endswith = ends with
    # __iendswith = ends with (case insensitive)
    # __exact = exact
    # __iexact = exact (case insensitive)
    # __regex = regular expression
    # __iregex = regular expression (case insensitive)
    # __gt = greater than
    # __lt = less than
    # __gte = greater than or equal to
    # __lte = less than or equal to

    filtered_data = Table1.objects.filter(integer_field__gte=5)  # Example filter
    filtered_data_1 = Table1.objects.filter(integer_field__contains=1)
    return render(request, 'queries.html', {
        'class': 'Basic',
        'type': 'Filter',
        'query1_name': 'Entries where integer_field is greater than or equal to 5',
        'query2_name': 'Entries where integer_field contains the digit 1',
        'query1': filtered_data,
        'query2': filtered_data_1
    })

# View to demonstrate getting specific records
def get_example(request):
    try:
        # Get first entry or None if it doesn't exist
        first_table1 = Table1.objects.get(id=1) if Table1.objects.filter(id=1).exists() else None
        # Get entry with highest integer_field value
        max_integer_entry = Table1.objects.get(integer_field=Table1.objects.aggregate(Max('integer_field'))['integer_field__max'])
        
        get_data = [first_table1] if first_table1 else []
        get_data_2 = [max_integer_entry] if max_integer_entry else []
        
        return render(request, 'queries.html', {
            'type': 'Get',
            'query1_name': 'First Table1 entry (ID=1) if exists',
            'query2_name': 'Table1 entry with highest integer_field value',
            'query1': get_data,
            'query2': get_data_2
        })
    except Exception as e:
        return render(request, 'queries.html', {
            'class': 'Basic',
            'type': 'Get',
            'query1_name': 'Error occurred',
            'query2_name': str(e),
            'query1': [],
            'query2': []
        })

# View to demonstrate exclude lookups
def exclude_example(request):
    excluded_data = Table1.objects.exclude(boolean_field=True)  # Example exclude
    return render(request, 'queries.html', {
        'class': 'Basic',
        'type': 'Exclude',
        'query1_name': 'Entries excluding boolean field = True',
        'query2_name': '',
        'query1': excluded_data,
        'query2': None
    })

# View to demonstrate ordering records
def order_by_example(request):
    # Order by integer_field ascending
    ordered_asc = Table1.objects.order_by('integer_field')
    # Order by integer_field descending
    ordered_desc = Table1.objects.order_by('-integer_field')
    
    return render(request, 'queries.html', {
        'class': 'Basic',
        'type': 'Order By',
        'query1_name': 'Table1 entries ordered by integer_field (ascending)',
        'query2_name': 'Table1 entries ordered by integer_field (descending)',
        'query1': ordered_asc,
        'query2': ordered_desc
    })

# View to demonstrate slicing data
def slice_example(request):
    sliced_data = Table1.objects.all()[:3]  # Example slice
    return render(request, 'queries.html', {
        'class': 'Basic',
        'type': 'Slice',
        'query1_name': 'First 3 entries from Table1',
        'query2_name': '',
        'query1': sliced_data,
        'query2': None
    })

def exists_example(request):
    # Check if any Table1 entries exist with boolean_field=True
    has_active = Table1.objects.filter(boolean_field=True).exists()
    # Check if any Table1 entries exist with integer_field > 10
    has_high_values = Table1.objects.filter(integer_field__gt=10).exists()
    
    # Prepare data for template
    value_names = ['Boolean Field Check', 'Integer Field Check']
    value_description = [
        'Checks if any records have "boolean_field=True"',
        'Checks if any records have "integer_field > 10"'
    ]
    values = [has_active, has_high_values]
    value_pairs = list(zip(value_names, value_description, values))
    
    return render(request, 'values_query.html', {
        'type': 'Basic Query Methods - Exists',
        'value_pairs': value_pairs,
    })

# View to demonstrate select_related for ForeignKey and OneToOne relationships
def select_related_example(request):
    # Fetch Table1 objects with their related ForeignKey and OneToOne objects
    select_related_data = Table1.objects.select_related('foreign_key', 'one_to_one')
    
    return render(request, 'queries.html', {
        'class': 'Advanced',
        'type': 'Select Related',
        'query1_name': 'Table1 with select_related ForeignKey and OneToOne relationships',
        'query2_name': '',
        'query1': select_related_data,
        'query2': None
    })

# View to demonstrate prefetching related data
def prefetch_related_example(request):
    # Retrieves all Table1 objects and, in a single additional query,
    # fetches all many-to-many relationships with Table3 for each object.
    # This way, accessing obj.many_to_many.all() does not generate extra queries (optimizes access).
    prefetch_data = Table1.objects.prefetch_related('many_to_many')
    
    return render(request, 'queries.html', {
        'class': 'Advanced',
        'type': 'Prefetch Related',
        'query1_name': 'Table1 with prefetched many-to-many relationships with Table3',
        'query2_name': '',
        'query1': prefetch_data,
        'query2': None
    })

# View to demonstrate F() expressions
def f_example(request):
    # Find entries where integer_field equals float_field (converted to int)
    # Note: This is a conceptual example - actual comparison depends on your data
    f_comparison = Table1.objects.filter(integer_field__isnull=False, float_field__isnull=False)
    
    # Example using F() in annotations - add 10 to integer_field
    f_annotation = Table1.objects.filter(integer_field__isnull=False).annotate(
        integer_plus_ten=F('integer_field') + 10
    )[:5]  # Limit to first 5 for display
    return render(request, 'queries.html', {
        'class': 'Advanced',
        'type': 'F() Expression',
        'query1_name': 'Table1 entries with non-null integer and float fields',
        'query2_name': 'Table1 entries with integer_field + 10 annotation (first 5)',
        'query1': f_comparison,
        'query2': f_annotation
    })

# View to demonstrate existence check
def Q_example(request):
    # Example of Q objects for complex queries
    q_data_or = Table1.objects.filter(Q(integer_field__gt=5) | Q(boolean_field=True))
    q_data_and = Table1.objects.filter(Q(integer_field__gt=5) & Q(boolean_field=True))

    return render(request, 'queries.html', {
        'class': 'Advanced',
        'type': 'Q',
        'query1_name': 'Table1 where integer_field > 5 OR boolean_field is True',
        'query2_name': 'Table1 where integer_field > 5 AND boolean_field is True',
        'query1': q_data_or,
        'query2': q_data_and
    })

# Agregate functions and counts
def query_values_example(request):
    # Count and basic aggregates
    count_data = Table1.objects.count()
    # max, min, avg for integer_field, float_field, date_field
    max_value = Table1.objects.aggregate(Max('integer_field'))
    min_integer = Table1.objects.aggregate(Min('integer_field'))
    min_float = Table1.objects.aggregate(Min('float_field'))
    min_date = Table1.objects.aggregate(Min('date_field'))
    avg_integer = Table1.objects.aggregate(Avg('integer_field'))
    avg_float = Table1.objects.aggregate(Avg('float_field'))
    
    # Count many-to-many relationships
    count_many_to_many = Table1.objects.annotate(num_table3=Count('many_to_many'))
    
    # Prepare data for template
    value_names = [
        'Total Count',
        'Max Integer Field',
        'Min Integer Field',
        'Min Float Field',
        'Min Date Field',
        'Average Integer Field',
        'Average Float Field',
        'ManyToMany Relations Count'
    ]
    # Descriptions for clarity
    value_description = [
        'Total number of records in Table1',
        'Maximum value of the integer_field in Table1',
        'Minimum value of the integer_field in Table1',
        'Minimum value of the float_field in Table1',
        'Earliest date in date_field in Table1',
        'Average value of the integer_field in Table1',
        'Average value of the float_field in Table1',
        'Count of related Table3 entries for each Table1 record'
    ]
    # Values corresponding to the above names and descriptions
    values = [
        count_data,
        max_value['integer_field__max'],
        min_integer['integer_field__min'],
        min_float['float_field__min'],
        min_date['date_field__min'],
        round(avg_integer['integer_field__avg'], 2) if avg_integer['integer_field__avg'] else None,
        round(avg_float['float_field__avg'], 2) if avg_float['float_field__avg'] else None,
        count_many_to_many
    ]
    # Combine names, descriptions, and values for template
    value_pairs = list(zip(value_names, value_description, values))
    return render(request, 'values_query.html', {
        'type': 'Aggregate Functions - Count, Max, Min, Avg & Annotations',
        'value_pairs': value_pairs,
    })

# Views for HTML examples
def html_modify(request):
    return render(request, 'html_modify.html')

# Example view to demonstrate various HTML template features
def html_example(request):
    return render(request, "html_example.html", {
        "title": "Welcome",
        "user": request.user,
        "date": timezone.now(),
        "number": 1234567,
        "products": [
            {"name": "Mouse", "price": 10.5, "stock": 5},
            {"name": "Keyboard", "price": 20.0, "stock": 0},
        ],
        "my_dict": {"key": "example value"},
        "my_list": ["Item 1", "Item 2"],
        "long_text": "This is a very very long text that will be truncated.",
        "range": range(1, 6),
    })

# Views for exporting data examples
def export_to_file(request): 
    return render(request, 'export_to_file.html')

# Export Table1 data to PDF
def export_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="table1.pdf"'

    # Create the PDF object, using the response object as its "file."
    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 50
    x = 50
    p.setFont("Helvetica", 10)
    # Title
    p.drawString(x, y, "Table1 Data Export")
    y -= 30

    registros = Table1.objects.all()
    # Table headers
    for reg in registros:
        # Create a line of text for each record
        texto = (
            f"ID: {reg.id} | "
            f"Int: {reg.integer_field} | "
            f"Float: {reg.float_field} | "
            f"Text: {reg.char_field} | "
            f"Bool: {reg.boolean_field} | "
            f"Date: {reg.datetime_field.strftime('%Y-%m-%d %H:%M')} |"
        )
        p.drawString(x, y, texto)
        y -= 15
        # If the y position is too low, create a new page
        if y < 60:
            p.showPage()
            y = height - 50
            p.setFont("Helvetica", 10)

    p.showPage()
    p.save()
    return response

# Export Table1 data to Excel
def export_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="table1.xlsx"'

    # Create an in-memory workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Table1"

    # Set up headers
    ws.append([
        "ID", "Integer", "Float", "Char", "Text",
        "Boolean", "Date", "Time", "Datetime"
    ])

    registros = Table1.objects.all()

    # Populate data rows
    for reg in registros:
        # Remove timezone info for Excel compatibility
        datetime_val = reg.datetime_field
        if datetime_val and is_aware(datetime_val):
            datetime_val = datetime_val.replace(tzinfo=None)
        else:
            datetime_val = reg.datetime_field

        # time field
        time_val = reg.time_field
        if time_val and is_aware(time_val):
            time_val = time_val.replace(tzinfo=None)
        else:
            time_val = reg.time_field

        # Append the row
        ws.append([
            reg.id,
            reg.integer_field,
            reg.float_field,
            reg.char_field,
            reg.text_field,
            reg.boolean_field,
            reg.date_field,
            time_val,
            datetime_val,
        ])

    # Adjust column widths
    width = [5, 10, 10, 15, 25, 20, 15, 12, 22]

    # Set the column widths
    for i, width2 in enumerate(width, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width2
    wb.save(response)
    return response

# View to handle email sending
@login_required
def email_send(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()
        
        try:
            # email content
            email_subject = f"{subject}"
            email_message = f"""
Name: {name}
Email: {email}

Message: {message}
"""
            
            # How to send the email
            send_mail(
                email_subject,
                email_message,
                settings.EMAIL_HOST_USER,
                [settings.CONTACT_EMAIL], # If want resend any email change this to "email" variable
                fail_silently=False,
            )
            
            messages.success(request, 'Your message has been sent successfully.')
            return redirect('email_send')
            
        except Exception as e:
            messages.error(request, f'Error sending message: {str(e)}')
            return render(request, 'email_send.html')
    
    return render(request, 'email_send.html')

# View to demonstrate custom template tags
def template_tags(request):
    data = Table1.objects.first()
    return render(request, 'template_tags.html', {'data': data})